from __future__ import annotations

import ast
import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

import cv2
import matplotlib
import numpy as np
from matplotlib import pyplot as plt
from matplotlib.patches import FancyArrowPatch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

matplotlib.use("Agg")

REPO = Path(r"F:\AAAzpj\code\Vert_detect")
ROOT = Path(r"C:\Users\35357\Desktop\final_exprements")
EXPORT = ROOT / "exports" / "idr_csv_export"
OUT = ROOT / "exports" / "journal_package_IDR_IRA_20260601_v4"
FIGURES = OUT / "figures"
TABLES = OUT / "tables"
SOURCE = OUT / "source_data"
IMAGE_DIR = REPO / "vert_all" / "val" / "images"
ANN_FILE = REPO / "vert_all" / "annotations" / "val.json"

CLASS_NAMES = [
    "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9",
    "T10", "T11", "T12", "L1", "L2", "L3", "L4", "L5", "S1",
]
FOV_GROUPS = ["Short (<=6)", "Medium (7-10)", "Long (>10)"]
REGION_GROUPS = ["Thoracic", "Lumbar-sacral", "Thoracolumbar"]
ABLATIONS = {
    "wo. CLIP2SAM": ("wo.CLIP2SAM", "crf_post"),
    "wo. SAM2CLIP": ("wo.SAM2CLIP", "crf_post"),
    "wo. RSCE": ("wo.RSCE", "crf_post"),
    "wo. CRF": ("Ours", "original"),
    "CLIP replacing MedCLIP": ("CLIP replacing MedCLIP", "crf_post"),
    "Ours": ("Ours", "crf_post"),
}
COLORS = {
    "YOLO26": "#4E79A7", "VertFound": "#F28E2B", "Ours": "#59A14F",
    "wo. CLIP2SAM": "#9C755F", "wo. SAM2CLIP": "#BAB0AC",
    "wo. RSCE": "#E15759", "wo. CRF": "#B07AA1",
    "CLIP replacing MedCLIP": "#76B7B2",
}


def parse_seq(value):
    if isinstance(value, list):
        return [int(x) for x in value]
    return [int(x) for x in ast.literal_eval(str(value))]


def read_csv(path):
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]) if rows else [])
        writer.writeheader()
        writer.writerows(rows)


def fov_group(seq):
    return FOV_GROUPS[0] if len(seq) <= 6 else FOV_GROUPS[1] if len(seq) <= 10 else FOV_GROUPS[2]


def region_group(seq):
    has_t = any(0 <= x <= 11 for x in seq)
    has_ls = any(12 <= x <= 17 for x in seq)
    return "Thoracolumbar" if has_t and has_ls else "Thoracic" if has_t else "Lumbar-sacral"


def make_dirs():
    for path in [OUT, FIGURES, TABLES, SOURCE]:
        path.mkdir(parents=True, exist_ok=True)


def savefig(fig, name):
    fig.savefig(FIGURES / f"{name}.png", dpi=300, bbox_inches="tight", pad_inches=0.12)
    fig.savefig(FIGURES / f"{name}.pdf", dpi=300, bbox_inches="tight", pad_inches=0.12)
    plt.close(fig)


def build_groups(per_image):
    groups = {}
    for row in per_image:
        seq = parse_seq(row["gt_seq"])
        groups[(row["experiment"], row["variant"], str(row["image_id"]))] = {
            "fov": fov_group(seq), "region": region_group(seq)
        }
    return groups


def aggregate_export(details, per_image, groups, experiment, variant, group_key, wanted):
    detail_rows = [r for r in details if r["experiment"] == experiment and r["variant"] == variant]
    image_rows = [r for r in per_image if r["experiment"] == experiment and r["variant"] == variant]
    out = []
    for group in wanted:
        d = [r for r in detail_rows if groups[(experiment, variant, str(r["image_id"]))][group_key] == group]
        p = [r for r in image_rows if groups[(experiment, variant, str(r["image_id"]))][group_key] == group]
        out.append({
            "Group": group, "N images": len(p), "N vertebrae": len(d),
            "IDR (%)": round(100 * sum(int(r["accepted_by_old_calc_IDR"]) for r in d) / len(d), 2) if d else 0.0,
            "IRA (%)": round(100 * sum(int(r["IRA_exact_sequence"]) for r in p) / len(p), 2) if p else 0.0,
        })
    return out


def read_vertfound_groups():
    seq_rows = read_csv(ROOT / "model_output" / "vertfound-master" /
                        "repro_vertfound_baseline_eval_val_4999" / "inference" / "eval_sequences_eval.csv")
    out = {}
    for row in seq_rows:
        seq = parse_seq(row["gt_seq"])
        out[str(row["image_id"])] = {"fov": fov_group(seq), "region": region_group(seq)}
    return out


def aggregate_vertfound(group_key, wanted):
    details = read_csv(EXPORT / "vertfound_baseline" / "vertfound_baseline_post_old_IDR_details.csv")
    per_image = read_csv(EXPORT / "vertfound_baseline" / "vertfound_baseline_post_old_IDR_per_image.csv")
    groups = read_vertfound_groups()
    out = []
    for group in wanted:
        d = [r for r in details if groups[str(r["image_id"])][group_key] == group]
        p = [r for r in per_image if groups[str(r["image_id"])][group_key] == group]
        out.append({
            "Group": group, "N images": len(p), "N vertebrae": len(d),
            "IDR (%)": round(100 * sum(int(r["accepted_by_old_calc_IDR"]) for r in d) / len(d), 2) if d else 0.0,
            "IRA (%)": round(100 * sum(int(r["IRA_exact_sequence"]) for r in p) / len(p), 2) if p else 0.0,
        })
    return out


def aggregate_yolo(group_key, wanted):
    rows = read_csv(EXPORT / "yolo26" / "eval_sequences_val.csv")
    for row in rows:
        seq = parse_seq(row["gt_seq"])
        row["fov"] = fov_group(seq)
        row["region"] = region_group(seq)
    out = []
    for group in wanted:
        selected = [r for r in rows if r[group_key] == group]
        total = sum(int(r["seq_len"]) for r in selected)
        out.append({
            "Group": group, "N images": len(selected), "N vertebrae": total,
            "IDR (%)": round(100 * sum(int(r["correct"]) for r in selected) / total, 2) if total else 0.0,
            "IRA (%)": round(100 * sum(int(r["IRA"]) for r in selected) / len(selected), 2) if selected else 0.0,
        })
    return out


def with_method(method, rows):
    return [{"Method": method, **row} for row in rows]


def overall(details, per_image, experiment, variant):
    d = [r for r in details if r["experiment"] == experiment and r["variant"] == variant]
    p = [r for r in per_image if r["experiment"] == experiment and r["variant"] == variant]
    return {
        "Method": experiment if experiment != "wo.RSCE" else "wo. RSCE",
        "N images": len(p), "N vertebrae": len(d),
        "IDR (%)": round(100 * sum(int(r["accepted_by_old_calc_IDR"]) for r in d) / len(d), 2),
        "IRA (%)": round(100 * sum(int(r["IRA_exact_sequence"]) for r in p) / len(p), 2),
    }


def radar(records, groups, title, name):
    methods = list(dict.fromkeys(row["Method"] for row in records))
    angles = np.linspace(0, 2 * np.pi, len(groups), endpoint=False).tolist()
    angles += angles[:1]
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.3), subplot_kw={"projection": "polar"})
    for ax, metric in zip(axes, ["IDR (%)", "IRA (%)"]):
        for method in methods:
            lookup = {r["Group"]: r[metric] for r in records if r["Method"] == method}
            values = [lookup[g] for g in groups] + [lookup[groups[0]]]
            ax.plot(angles, values, lw=2.2, marker="o", ms=4, color=COLORS[method], label=method)
            ax.fill(angles, values, color=COLORS[method], alpha=0.05)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(groups, fontsize=9)
        ax.set_ylim(0, 100)
        ax.set_title(metric, fontweight="bold", pad=14)
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncol=3, frameon=False)
    fig.suptitle(title, fontsize=16, fontweight="bold")
    fig.subplots_adjust(bottom=0.18, top=0.84, wspace=0.35)
    savefig(fig, name)


def metric_bars(rows, title, name):
    fig, ax = plt.subplots(figsize=(7.4, 4.6))
    x = np.arange(len(rows))
    width = 0.34
    ax.bar(x - width / 2, [r["IDR (%)"] for r in rows], width, label="IDR", color="#4E79A7")
    ax.bar(x + width / 2, [r["IRA (%)"] for r in rows], width, label="IRA", color="#F28E2B")
    ax.set_xticks(x)
    ax.set_xticklabels([r["Method"] for r in rows])
    ax.set_ylim(0, 105)
    ax.set_ylabel("Rate (%)")
    ax.set_title(title, fontsize=15, fontweight="bold")
    ax.legend(frameon=False)
    ax.grid(axis="y", alpha=0.24)
    ax.spines[["top", "right"]].set_visible(False)
    for idx, row in enumerate(rows):
        ax.text(idx - width / 2, row["IDR (%)"] + 1.2, f'{row["IDR (%)"]:.2f}', ha="center", fontsize=8)
        ax.text(idx + width / 2, row["IRA (%)"] + 1.2, f'{row["IRA (%)"]:.2f}', ha="center", fontsize=8)
    savefig(fig, name)


def plot_chords(details):
    def transitions(variant):
        rows = [r for r in details if r["experiment"] == "Ours" and r["variant"] == variant]
        return Counter((int(r["gt_category_id"]) - 1, int(r["pred_category_id"]) - 1)
                       for r in rows if r["gt_category_id"] != r["pred_category_id"])
    mats = [transitions("original"), transitions("crf_post")]
    vmax = max(max(x.values(), default=1) for x in mats)
    n = len(CLASS_NAMES)
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False)
    points = {i: np.array([math.cos(a), math.sin(a)]) for i, a in enumerate(angles)}
    cmap = plt.get_cmap("turbo")
    fig, axes = plt.subplots(1, 2, figsize=(12, 6))
    for ax, trans, title in zip(axes, mats, ["(a) wo. CRF", "(b) Ours"]):
        ax.axis("off")
        ax.set_aspect("equal")
        ax.add_artist(plt.Circle((0, 0), 1, fill=False, color="#D8D8D8"))
        for i, label in enumerate(CLASS_NAMES):
            p = points[i]
            ax.scatter(*p, s=45, color=cmap(i / (n - 1)), edgecolor="white", zorder=4)
            ax.text(*(p * 1.15), label, ha="center", va="center", fontsize=8)
        for (src, dst), count in trans.most_common(45):
            if not (0 <= src < n and 0 <= dst < n):
                continue
            rad = 0.24 if (dst - src) % n < n / 2 else -0.24
            ax.add_patch(FancyArrowPatch(points[src] * .96, points[dst] * .96,
                         connectionstyle=f"arc3,rad={rad}", arrowstyle="-|>",
                         mutation_scale=6 + 7 * count / vmax, lw=.45 + 3 * count / vmax,
                         color=cmap(src / (n - 1)), alpha=.38))
        ax.set_xlim(-1.32, 1.32)
        ax.set_ylim(-1.28, 1.28)
        ax.set_title(title, fontweight="bold")
    fig.suptitle("Misclassification Transition Chords", fontsize=16, fontweight="bold")
    fig.text(.5, .03, "Arc direction: ground truth -> prediction; thicker arcs indicate more errors.",
             ha="center", color="#666666")
    savefig(fig, "05_crf_transition_chords")


def plot_stars(details):
    prepared = []
    vmax = 1
    for label, (experiment, variant) in ABLATIONS.items():
        rows = [r for r in details if r["experiment"] == experiment and r["variant"] == variant]
        errors = Counter(int(r["gt_category_id"]) - 1 for r in rows if r["gt_category_id"] != r["pred_category_id"])
        values = [errors[i] for i in range(len(CLASS_NAMES))]
        vmax = max(vmax, max(values, default=0))
        prepared.append((label, values))
    theta = np.linspace(0, 2 * np.pi, len(CLASS_NAMES), endpoint=False)
    fig, axes = plt.subplots(2, 3, figsize=(15, 8.5), subplot_kw={"projection": "polar"})
    colors = [plt.get_cmap("turbo")(i / (len(CLASS_NAMES) - 1)) for i in range(len(CLASS_NAMES))]
    for ax, (label, values) in zip(axes.flat, prepared):
        ax.bar(theta, values, width=2 * np.pi / len(CLASS_NAMES) * .72, color=colors, alpha=.9)
        ax.set_xticks(theta)
        ax.set_xticklabels(CLASS_NAMES, fontsize=7)
        ax.set_yticklabels([])
        ax.set_ylim(0, vmax)
        ax.grid(alpha=.3)
        ax.set_title(label, fontsize=12, fontweight="bold", pad=12)
    fig.suptitle("Ablation Error Visualization by Confusion Stars", fontsize=17, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, .95], h_pad=1.5)
    savefig(fig, "06_ablation_confusion_stars")


def load_annotations():
    data = json.loads(ANN_FILE.read_text(encoding="utf-8"))
    images = {str(x["id"]): x for x in data["images"]}
    by_name = {Path(x["file_name"]).stem: str(x["id"]) for x in data["images"]}
    anns = defaultdict(list)
    for ann in data["annotations"]:
        anns[str(ann["image_id"])].append(ann)
    for key in anns:
        anns[key].sort(key=lambda x: (x["bbox"][1], x["bbox"][0]))
    return images, by_name, anns


def draw_case(image, anns, seq, title):
    canvas = image.copy()
    cmap = plt.get_cmap("turbo")
    for ann, label in zip(anns, seq):
        x, y, w, h = [int(round(v)) for v in ann["bbox"]]
        rgb = cmap(int(label) / (len(CLASS_NAMES) - 1))[:3]
        color = tuple(int(v * 255) for v in rgb[::-1])
        cv2.rectangle(canvas, (x, y), (x + w, y + h), color, 2, cv2.LINE_AA)
        cv2.putText(canvas, CLASS_NAMES[int(label)], (x + 2, y + 17),
                    cv2.FONT_HERSHEY_SIMPLEX, .53, color, 2, cv2.LINE_AA)
    cv2.putText(canvas, title, (12, 30), cv2.FONT_HERSHEY_SIMPLEX, .82, (255, 255, 255), 2, cv2.LINE_AA)
    return cv2.cvtColor(canvas, cv2.COLOR_BGR2RGB)


def qualitative_panels(per_image):
    images, by_name, anns = load_annotations()
    rows = {(r["variant"], str(r["image_id"])): r for r in per_image if r["experiment"] == "Ours"}
    selected = {
        "Thoracic": ["DR_204_s1_e12", "DR_102_s1_e12", "DR_426_s3_e12"],
        "Lumbar-sacral": ["gq346", "gs30", "gs31"],
    }
    for region, ids in selected.items():
        fig, axes = plt.subplots(3, 3, figsize=(14, 13))
        for row_idx, image_id in enumerate(ids):
            before = rows[("original", image_id)]
            after = rows[("crf_post", image_id)]
            coco_id = image_id if image_id in images else by_name[image_id]
            file_name = Path(images[coco_id]["file_name"]).name
            image = cv2.imread(str(IMAGE_DIR / file_name), cv2.IMREAD_COLOR)
            panels = [
                draw_case(image, anns[coco_id], parse_seq(before["gt_seq"]), "Ground Truth"),
                draw_case(image, anns[coco_id], parse_seq(before["pred_seq"]), "wo. CRF"),
                draw_case(image, anns[coco_id], parse_seq(after["pred_seq"]), "Ours"),
            ]
            for col_idx, panel in enumerate(panels):
                axes[row_idx, col_idx].imshow(panel)
                axes[row_idx, col_idx].axis("off")
                if row_idx == 0:
                    axes[row_idx, col_idx].set_title(["Ground Truth", "wo. CRF", "Ours"][col_idx],
                                                     fontsize=14, fontweight="bold")
            axes[row_idx, 0].text(.01, -.05, f"Case {image_id}", transform=axes[row_idx, 0].transAxes, fontsize=10)
        fig.suptitle(f"Qualitative Examples: {region}", fontsize=17, fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, .97])
        suffix = "thoracic" if region == "Thoracic" else "lumbar_sacral"
        savefig(fig, f"07_qualitative_examples_{suffix}")


def write_workbook(table_specs):
    wb = Workbook()
    wb.remove(wb.active)
    for name, title, rows in table_specs:
        ws = wb.create_sheet(name)
        ws.append([title])
        ws.append(list(rows[0]))
        for row in rows:
            ws.append(list(row.values()))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
        ws["A1"].font = Font(bold=True)
        for cell in ws[2]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        for col in "ABCDE":
            ws.column_dimensions[col].width = 18
    wb.save(TABLES / "journal_tables_IDR_IRA.xlsx")


def main():
    make_dirs()
    details = read_csv(EXPORT / "idr_reconstructed_details_all.csv")
    per_image = read_csv(EXPORT / "idr_reconstructed_per_image_all.csv")
    groups = build_groups(per_image)
    ours_fov = aggregate_export(details, per_image, groups, "Ours", "crf_post", "fov", FOV_GROUPS)
    ours_region = aggregate_export(details, per_image, groups, "Ours", "crf_post", "region", REGION_GROUPS)
    comparison_fov = with_method("YOLO26", aggregate_yolo("fov", FOV_GROUPS))
    comparison_fov += with_method("VertFound", aggregate_vertfound("fov", FOV_GROUPS))
    comparison_fov += with_method("Ours", ours_fov)
    comparison_region = with_method("YOLO26", aggregate_yolo("region", REGION_GROUPS))
    comparison_region += with_method("VertFound", aggregate_vertfound("region", REGION_GROUPS))
    comparison_region += with_method("Ours", ours_region)
    ablation_fov = []
    ablation_region = []
    for label, (experiment, variant) in ABLATIONS.items():
        ablation_fov += with_method(label, aggregate_export(details, per_image, groups, experiment, variant, "fov", FOV_GROUPS))
        ablation_region += with_method(label, aggregate_export(details, per_image, groups, experiment, variant, "region", REGION_GROUPS))
    specs = [
        ("Table_S1", "Comparison under different field-of-view lengths", comparison_fov),
        ("Table_S2", "Ablation study under different field-of-view lengths", ablation_fov),
        ("Table_S3", "Comparison under different anatomical coverage regions", comparison_region),
        ("Table_S4", "Ablation study under different anatomical coverage regions", ablation_region),
    ]
    for idx, (_, _, rows) in enumerate(specs, 1):
        write_csv(SOURCE / f"table_s{idx}.csv", rows)
    write_workbook(specs)
    radar(comparison_fov, FOV_GROUPS, "Method Comparison across FoV Lengths", "01_fov_method_comparison")
    radar(comparison_region, REGION_GROUPS, "Method Comparison across Anatomical Regions", "02_region_method_comparison")
    crf_rows = [overall(details, per_image, "Ours", "original"), overall(details, per_image, "Ours", "crf_post")]
    crf_rows[0]["Method"], crf_rows[1]["Method"] = "wo. CRF", "Ours"
    rsce_rows = [overall(details, per_image, "wo.RSCE", "crf_post"), overall(details, per_image, "Ours", "crf_post")]
    rsce_rows[0]["Method"], rsce_rows[1]["Method"] = "wo. RSCE", "Ours"
    write_csv(SOURCE / "overall_crf_IDR_IRA.csv", crf_rows)
    write_csv(SOURCE / "overall_rsce_IDR_IRA.csv", rsce_rows)
    metric_bars(crf_rows, "Effect of the CRF Module", "03_crf_module_IDR_IRA")
    metric_bars(rsce_rows, "Effect of RSCE", "04_rsce_IDR_IRA")
    plot_chords(details)
    plot_stars(details)
    qualitative_panels(per_image)
    (OUT / "README.md").write_text(
        "# Journal figures: IDR and IRA only\n\n"
        "This package uses only the requested IDR and IRA metrics. It intentionally excludes "
        "the deprecated alternative metric names from figures, tables, filenames, and notes.\n\n"
        "Figures 01-02 retain the method-comparison views. Ablation radar figures are omitted. "
        "Figure 04 shows the RSCE comparison. Figure 06 has no numeric footer under the stars. "
        "Figures 07 split qualitative cases into one thoracic panel and one lumbar-sacral panel.\n",
        encoding="utf-8",
    )
    print(OUT)


if __name__ == "__main__":
    main()
